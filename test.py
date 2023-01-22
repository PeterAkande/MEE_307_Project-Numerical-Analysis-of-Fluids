import math
import os.path

import matplotlib.pyplot as plt
import openpyxl

ENV_TYPE = 'horizontal'
ACCELERATION_DUE_GRAVITY = 0.01
VELOCITY = [0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5]
DIAMETERS = [0.00159, 0.00318, 0.00477, 0.00636, 0.00795, 0.00954, 0.01113, 0.01272, 0.01431, 0.0159]
LENGTHS = [0.2, 0.4, 0.6, 0.8, 1.0, 1.2, 1.4, 1.6, 1.8, 2.0]
PIPE_ROUGHNESS = 0.03
THERMAL_CONDUCTIVITY = 401


def calculate_reynolds_number(density: float, diameter: float, velocity: float, dynamic_viscosity: float) -> float:
    # This function calculates the reynolds number

    reynolds_number = (density * diameter * velocity) / dynamic_viscosity

    return reynolds_number


def calculate_frictional_factor_for_laminar_flow(reynold_number: float) -> float:
    return 64 / reynold_number


def calculate_frictional_factor_for_turbulent(reynold_number: float, pipe_roughness: float, diameter: float) -> float:
    return 2 * ((8 / reynold_number) ** 12 + (
            (2.457 * math.log((0.27 * pipe_roughness / diameter) + (7 / reynold_number) ** 0.9)) ** 16 + (
            37530 / reynold_number) ** 16) ** (-3 / 2)) ** (
                   1 / 12)


def get_frictional_factor(reynold_number: float, pipe_roughness: float, is_laminar: bool, diameter: float) -> float:
    if is_laminar:
        frictional_factor = calculate_frictional_factor_for_laminar_flow(reynold_number)
    else:
        frictional_factor = calculate_frictional_factor_for_turbulent(reynold_number=reynold_number,
                                                                      pipe_roughness=pipe_roughness, diameter=diameter)

    return frictional_factor


def calculate_head_loss(friction_factor: float, pipe_length: float, diameter: float, velocity: float):
    head_loss = (friction_factor * pipe_length * velocity ** 2) / (diameter * 2 * ACCELERATION_DUE_GRAVITY)
    return head_loss


def calculate_prandtl_number(dynamic_viscosity: float, specific_heat_capacity: float, conductivity: float) -> float:
    prandtl_number = (dynamic_viscosity * specific_heat_capacity) / conductivity

    return prandtl_number


def calculate_pressure(density: float, head_loss: float) -> float:
    return -density * ACCELERATION_DUE_GRAVITY * head_loss


def calculate_coefficient_of_heat_transfer(reynold_number: float, diameter: float, prandtl_number: float,
                                           conductivity: float) -> float:
    return 0.023 * (reynold_number ** 0.8) * (prandtl_number ** 0.4) * conductivity / diameter


def get_values(specific_heat_capacity: float, dynamic_viscosity: float, density: float) -> dict:
    # This function gets the various values and does the operation for each fluid

    reynolds_number_values = []  # This list would store the reynold number for the fluid in what ever case
    head_loss_values = []  # This list would store the head loss for the fluid in varuous cases
    coefficient_of_heat_transfer_values = []
    frictional_factors = []

    # Get the values that would be used for all fluids
    # specific_heat_capacity = float(input('Enter specific heat capacity: '))
    # dynamic_viscosity = float(input('Enter the dynamic viscosity: '))
    # density = float(input('Enter the density of the fluid: '))

    for length, diameter, velocity in zip(LENGTHS, DIAMETERS, VELOCITY):
        # Calculate the reynold's number.
        reynold_number = calculate_reynolds_number(diameter=diameter, density=density, velocity=velocity,
                                                   dynamic_viscosity=dynamic_viscosity)

        # Get if the fluid is laminar or not
        is_laminar = reynold_number < 2000

        # Get the prandtl number
        prandtl_number = calculate_prandtl_number(specific_heat_capacity=specific_heat_capacity,
                                                  conductivity=THERMAL_CONDUCTIVITY,
                                                  dynamic_viscosity=dynamic_viscosity)

        # Get the frictional factor
        frictional_factor = get_frictional_factor(reynold_number=reynold_number, pipe_roughness=PIPE_ROUGHNESS,
                                                  is_laminar=is_laminar, diameter=diameter)

        # Get the head loss
        head_loss = calculate_head_loss(friction_factor=frictional_factor, pipe_length=length, diameter=diameter,
                                        velocity=velocity)

        # Get the pressure
        pressure = calculate_pressure(density=density, head_loss=head_loss)

        # Get the coefficient of heat transfer.
        coefficient_of_heat_transfer = calculate_coefficient_of_heat_transfer(reynold_number=reynold_number,
                                                                              diameter=diameter,
                                                                              prandtl_number=prandtl_number,
                                                                              conductivity=THERMAL_CONDUCTIVITY)

        frictional_factors.append(frictional_factor)
        coefficient_of_heat_transfer_values.append(coefficient_of_heat_transfer)
        head_loss_values.append(head_loss)
        reynolds_number_values.append(reynold_number)

    return {
        'head_loss': head_loss_values,
        'frictional_factor': frictional_factors,
        'heat_transfer_coefficient': coefficient_of_heat_transfer_values,
        'reynolds_number': reynolds_number_values,
        'velocity': VELOCITY,
        'diameter': DIAMETERS,
    }


def plot_graph(x: list, y: list, x_axis_label, y_axis_label, plot_title):
    plt.plot(x, y, color='r')
    plt.xlabel(x_axis_label)
    plt.ylabel(y_axis_label)
    # setting the title of the plot
    plt.title(plot_title)
    plt.show()


# sub plots
def add_subplot_graph(axis, row: int, col: int, x: list, y: list, x_axis_label: str, y_axis_label: str,
                      plot_title: str, color: str):
    axis[row, col].plot(x, y, color=color)
    axis.set_title(plot_title)

    axis.xlabel(x_axis_label)
    axis.ylabel(y_axis_label)


def create_excel_sheet(fluids_values_dict: dict, fluid_name: str):
    """
    This would create an excel file from a dictionary.
    It is kind of hard coded in this case.
    It would create multiple excel sheets for each fluids.


    So, the keys of the fluids values dict would be the columns
    so for a fluid value say R407A:
    {
        head_loss: [] # The numbers in the list would be the column values,
        reynolds_number: [] # Same as this one
    }

    so each would be stored in the directory "cwd/generated/horizontal/{fluid_name}
    """
    assert (len(fluids_values_dict) > 0)

    cwd = os.getcwd()  # Get the current working directory
    path_to_file = os.path.join(cwd, 'generated', 'excel_sheets', ENV_TYPE, f"{fluid_name}_{ENV_TYPE}.xlsx")

    wb = openpyxl.Workbook()  # Create a workbook
    sheet = wb.active

    sheet.title = f"Values for Fluid {fluid_name}"

    number_of_columns = len(fluids_values_dict) + 1  # Include the length

    # Now add the headers
    headers = ['length'] + list(fluids_values_dict.keys())

    # Write the headers to the Excel sheet
    for i in range(number_of_columns):
        cell = sheet.cell(row=1, column=i + 1)
        cell.value = headers[i]

    for column in range(len(headers)):
        header = headers[column]

        if header == 'length':
            values = LENGTHS
        else:
            values = fluids_values_dict[header]

        # Now loop through the values and add them to the sheet.
        for _row in range(len(values)):
            # The row starts from row 2
            value = values[_row]
            cell = sheet.cell(row=_row + 2, column=column + 1)
            cell.value = value

    wb.save(filename=path_to_file)


if __name__ == '__main__':
    # Run some code

    fluids = ['R407a', 'R245fa', 'R1234ze', 'R1234yf', 'Water', 'Ammonia', 'R134a', 'Propane', 'R600a', 'R407c']
    colors = ['r', 'g', 'b', 'c', 'm', 'y', 'k', 'pink', 'chartreuse', 'burlywood']
    print('Enter fluid names. Press enter to stop recording')

    print(f'Calculating for {",".join(fluids)}')
    fluids_values = {
        fluids[0]: {
            'shc': 1520,
            'density': 1145.1,
            'viscosity': 0.000151
        }, fluids[1]: {
            'shc': 1322,
            'density': 1339,
            'viscosity': 0.000401
        }, fluids[2]: {
            'shc': 1386,
            'density': 1163.1,
            'viscosity': 0.000199
        }, fluids[3]: {
            'shc': 1392,
            'density': 1092,
            'viscosity': 0.000154
        }, fluids[4]: {
            'shc': 4187,
            'density': 1000,
            'viscosity': 0.000895
        }, fluids[5]: {
            'shc': 4744,
            'density': 696,
            'viscosity': 0.000255
        }, fluids[6]: {
            'shc': 1430,
            'density': 1207.2,
            'viscosity': 0.000181
        }, fluids[7]: {
            'shc': 1630,
            'density': 495,
            'viscosity': 0.00011
        }, fluids[8]: {
            'shc': 2430,
            'density': 551,
            'viscosity': 0.000151
        }, fluids[9]: {
            'shc': 1540,
            'density': 1134,
            'viscosity': 0.000154
        },
    }

    fluids_tables = {}

    for fluid in fluids:
        print(f'<------------Calculating for {fluid}-------------->')
        fluid_value = get_values(fluids_values[fluid]['shc'], fluids_values[fluid]['viscosity'],
                                 fluids_values[fluid]['density'], )
        fluids_tables[fluid] = fluid_value
        create_excel_sheet(fluid_value, fluid)

    print('Plotting graphs >>>>>>>>>>>> Loading >>>>>>>>>>>>>>>>>>>')

    graphs_details = [['head_loss', 'reynolds_number'], ['heat_transfer_coefficient', 'reynolds_number'],
                      ['frictional_factor', 'reynolds_number'], ['reynolds_number', 'diameter'],
                      ['reynolds_number', 'velocity']]

    for row in range(len(graphs_details)):

        graph_detail = graphs_details[row]
        y_axis = graph_detail[0]
        x_axis = graph_detail[1]

        for specific_graph_number in range(len(fluids)):
            # Plot all the graphs on a canvas
            plt.plot(fluids_tables[fluids[specific_graph_number]][x_axis],
                     fluids_tables[fluids[specific_graph_number]][y_axis],
                     color=colors[specific_graph_number])

        plt.xlabel(x_axis)
        plt.ylabel(y_axis)
        plt.title(f'graph of {y_axis} against {x_axis}')
        plt.legend(fluids)

        cwd = os.getcwd()
        path_to_file = os.path.join(cwd, 'generated', 'images', ENV_TYPE, f'{y_axis} against {x_axis}_{ENV_TYPE}.png')

        plt.savefig(path_to_file)
        # plt.show() # Plot the graph and show a window
